import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.colors import Color
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os


def individual_marksheet_generater(positive_score,negative_score): # Defining function for generating marksheet
    # Thin Border
    main_path=os.getcwd()
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )
    master_roll=[] #List to store rolll number in master list

    
    
   
    input_path_master_roll=os.path.join(main_path,'input','master_roll.csv')

    with open(input_path_master_roll) as file:
        reader=csv.reader(file)
        next(reader)
        for row in reader:
            master_roll.append(row[0].upper())
    output_details={}  #Dictionary whose value is list ==> [name, positive number, negative number, total marks, score]
    right_ans=[]
    stud_ans={} #Dictionary of students answer
    check=0
    minus=int(negative_score)
    plus=int(positive_score)
    response_details={}  #Dictionary for storing response details
    responded_roll=[]

    
    
    with open(os.path.join(main_path,'input','responses.csv'),'r') as file:
        reader=csv.reader(file)
        next(reader)
        for row in reader:
            rollnum=row[6].upper()
            responded_roll.append(rollnum)  #Appending roll numbers of students who attended examination to responded_roll list 
            right=0
            wrong=0
            name=row[3]
            score=row[2]
            responses=row[7:]
            stud_ans[rollnum]=responses
            if(check==0):
                right_ans.extend(responses)
            check+=1
            for i in range(len(responses)):
                if(responses[i]==right_ans[i]):
                    right+=1
                elif(responses[i]==''):
                    pass
                else:
                    wrong+=1
            positive=plus*right
            negative=minus*wrong
            total=positive + negative
            output_details[rollnum]=[name,right,wrong,total,score]

    try:
        os.mkdir('marksheets')
    except:
        pass
    overall_mark=len(right_ans)*plus

    

    for rollnum in master_roll: #Iterating throughout roll number
        output_path=os.path.join(main_path,'marksheets')
        os.chdir(output_path)

        if(rollnum not in responded_roll):    #If the student is absent
            wb.save(rollnum + '.xlsx')
        else:
            try:
                wb=Workbook() #For individiual result sheet
                sheet=wb.active
                img=Image(os.path.join(main_path,'pic.png'))
                sheet.add_image(img,'A1')          #Adding Logo and heading

                #Adding student details and exam details
                sheet['A8']='Name:'
                sheet['A8'].alignment = Alignment(horizontal='right')
                sheet['A9']='Roll Number:'
                sheet['A9'].alignment = Alignment(horizontal='right')
                sheet['D8']='Exam:'
                sheet['D8'].alignment = Alignment(horizontal='right')
                sheet['B8']=output_details[rollnum][0]
                sheet['B8'].alignment = Alignment(horizontal='left')
                sheet['B9']=rollnum
                sheet['B9'].alignment = Alignment(horizontal='left')
                sheet['E8']='quiz'
                sheet['E8'].alignment = Alignment(horizontal='left')

                #Overall Marking of Examination
                sheet['B11']='Right'
                sheet['C11']='Wrong'
                sheet['D11']='Not Attempt'
                sheet['E11']='Max'
                sheet['A12']='No.'
                sheet['A13']='Marking'
                sheet['A14']='Total'
                sheet['B12']=output_details[rollnum][1] #Correct answers
                sheet['C12']=output_details[rollnum][2] #Wrong Answers
                sheet['E12']=len(right_ans)
                sheet['B13']=plus
                sheet['C13']=minus
                sheet['D13']=0
                sheet['B14']=int(output_details[rollnum][1])*plus
                sheet['C14']=int(output_details[rollnum][2])*minus
                score=int(output_details[rollnum][1])*plus + int(output_details[rollnum][2])*minus
                sheet['E14']=str(output_details[rollnum][3]) + '/' + str(overall_mark)
                
                not_attempt=len(right_ans)-int(output_details[rollnum][1])-int(output_details[rollnum][2]) #Not attempted
                sheet['D12']=not_attempt
                sheet.row_dimensions[1].height = 20
                for col in 'ABCDE':
                    sheet.column_dimensions[col].width =20
                    for row in range(8,10):
                        sheet[col+str(row)].font=Font(name='Century',size=12)

                #Providing Style and Alignment
                for col in 'ABCDE':
                    for row in ['11','12','13','14']:
                        sheet[col+row].alignment = Alignment(horizontal='center')
                        sheet[col+row].font=Font( name='Century', size=12)
                        sheet[col+row].border=thin_border

                sheet['E8'].font=Font(bold=True,name='Century',size=12 )
                sheet['B9'].font = Font(bold=True,name='Century',size=12)
                sheet['B8'].font = Font(bold=True, name='Century', size=12)
                for col in 'BCDE':
                    sheet[col+'11'].font=Font(bold=True,name='Century',size=12)
                    
                for row in ['12','13','14']:
                    sheet['A'+row].font=Font( bold=True, name='Century', size=12)
                    sheet['B'+row].font=Font( name='Century', size=12,color='FF008000')
                    sheet['C'+row].font=Font( name='Century', size=12,color='FFFF0000')
                    sheet['D'+row].font=Font( name='Century', size=12)
                    sheet['E'+row].font=Font( name='Century', size=12)

                sheet['E14'].font=Font(name='Century', size=12,color='FF0000FF') 

            #Providing heading for student answer and correct answer
                sheet['A17']='Student Ans'
                sheet['A17'].font=Font(bold=True,name='Century', size=12)
                sheet['A17'].border=thin_border

                sheet['B17']='Correct Ans'
                sheet['B17'].font=Font(bold=True,name='Century', size=12)
                sheet['B17'].border=thin_border

                sheet['D17']='Student Ans'
                sheet['D17'].font=Font(bold=True,name='Century', size=12)
                sheet['D17'].border=thin_border

                sheet['E17']='Correct Ans'
                sheet['E17'].font=Font(bold=True,name='Century', size=12)
                sheet['E17'].border=thin_border

                for row in range(18,len(right_ans)+18):  #Providing different colour to right , wrong answers and correct answer
                    i=row-18
                    nrow=row
                    if(i<len(right_ans)-5):
                        left='A'
                        right='B'
                        nrow=row
                    else:
                        left='D'
                        right='E'
                        nrow=row-(len(right_ans)-5)
                    sheet[left+str(nrow)]=stud_ans[rollnum][i]
                    
                    if(stud_ans[rollnum][i]==right_ans[i]):
                        sheet[left+str(nrow)].font=Font(name='Century', size=12,color='FF008000') #Green for right 
                    else:
                        sheet[left+str(nrow)].font=Font(name='Century', size=12,color='FFFF0000') #Red for wrong
                    sheet[right+str(nrow)]=right_ans[i]
                    sheet[right+str(nrow)].font=Font(name='Century', size=12,color='FF0000FF') #Blue for choosed correct answer 

                    
                    #Adjusting alignment
                    sheet[right+str(nrow)].alignment = Alignment(horizontal='center')
                    sheet[left+str(nrow)].alignment = Alignment(horizontal='center')
                    sheet[left+str(nrow)].border=thin_border  #Applying border
                    sheet[right+str(nrow)].border=thin_border
                 
                os.chdir(output_path)
                wb.save(rollnum + '.xlsx')
                

            except:
                print('No roll number with Answer is present for',rollnum)   #If present and no answer
    os.chdir('../')

def concise_marksheet_generater(right_score,wrong_score):
    try:
        os.mkdir('marksheets')
    except:
        pass
    wrkbook=Workbook() #For Concise marksheet
    ws=wrkbook.active
    #Headers of Concise Marksheet
    ws.append(['Timestamp','Email address','Google_Score','Name','IITP webmail','Phone (10 digit only)','Roll Number'])
   
    main_path=os.getcwd()
    master_roll_list=[]
    responded_roll=[]
    answer_key=[]
    stud_ans={}
    output_details={}

    right_score=int(right_score)
    wrong_score=int(wrong_score)

    with open(os.path.join(main_path,'input','master_roll.csv')) as file:
        reader=csv.reader(file)
        next(reader)
        for row in reader:
            master_roll_list.append(row[0])

    with open(os.path.join(main_path,'input','responses.csv')) as file:
        reader=csv.reader(file)
        next(reader)
        for row in reader:
            roll_num=row[6].upper()
            responded_roll.append(roll_num)
            responses=row[7:]
            stud_ans[roll_num]=responses
            right=0
            wrong=0
            if(roll_num=='ANSWER'):
                answer_key=stud_ans[roll_num]
            for i in range(len(answer_key)):
                if(stud_ans[roll_num][i]==answer_key[i]):
                    right+=1
                elif(stud_ans[roll_num][i]==''):
                    pass
                else:
                    wrong+=1
            total_question=len(answer_key)
            score=right*right_score + wrong*wrong_score
            unattempted=total_question-right-wrong

            output_details[roll_num]=[score,right,wrong,unattempted]
            row[6]=roll_num
            ws.append(row)
    ws.cell(row=1, column=len(answer_key)+8).value='statusAns'
    os.chdir(os.path.join(os.getcwd(),'marksheets'))        
    wrkbook.save('concise_marksheet.xlsx')
    wb = load_workbook('concise_marksheet.xlsx')
    sheet = wb.active
    sheet.insert_cols(7) #Inserting column for adding score after considering negative marking
    sheet.cell(row=1, column=7).value='Score_After_Negative'
    for row in range(2,len(responded_roll)+2) :
        sheet.cell(row=row, column=7).value=str(output_details[responded_roll[row-2]][0]) + '/' + str(len(answer_key)*right_score)
        p_score=output_details[responded_roll[row-2]][1]
        n_score=output_details[responded_roll[row-2]][2]
        n_attempt=output_details[responded_roll[row-2]][3]
        statusAns=[p_score,n_score,n_attempt] #Status of answer=[correct ans, wrong ans, not attempted]
        sheet.cell(row=row, column=len(answer_key)+9).value=str(statusAns)

    row=len(responded_roll)+2

    for roll in master_roll_list:
        if(roll not in responded_roll):
            sheet.cell(row=row, column=8).value=roll
    wb.save('concise_marksheet.xlsx')
    os.chdir('../')


def email_send(subject="Examination Marks",body="Here is your Marks",sender_email="vishwaranjangopi@gmail.com",sender_pass="Gopi@100"):
    
    main_path=os.getcwd()
    roll_list=[]
    email_list=[]
    with open(os.path.join(main_path,'input','responses.csv'),'r') as file:
        reader=csv.reader(file)
        next(reader)
        
        for row in reader:
            roll_list.append(row[6].upper())
            email_list.append(row[1])
    
    fromaddr = sender_email

    subject=subject
     # creates SMTP session
    s = smtplib.SMTP('smtp.gmail.com', 587)
    # start TLS for security
    s.ehlo()
    s.starttls()
    s.ehlo()
    
    # Authentication
    s.login(fromaddr, sender_pass)
    for i in range(len(roll_list)):
        toaddr = email_list[i]
        # instance of MIMEMultipart
        msg = MIMEMultipart()
        # attach the body with the msg instance
        msg.attach(MIMEText(body, 'plain'))

        # storing the senders email address  
        msg['From'] = fromaddr
        # storing the receivers email address 
        msg['To'] = toaddr
        # storing the subject 
        msg['Subject'] = subject 
        # string to store the body of the mail
        body = body

        # open the file to be sent 
        filename = roll_list[i]+'.xlsx'
        attachment = open(os.path.join(main_path,"marksheets",filename), "rb")
        # instance of MIMEBase and named as p
        p = MIMEBase('application', 'octet-stream')
        # To change the payload into encoded form
        p.set_payload((attachment).read())
        # encode into base64
        encoders.encode_base64(p) 
        p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
        # attach the instance 'p' to instance 'msg'
        msg.attach(p)
        # Converts the Multipart msg into a string
        text = msg.as_string()
        # sending the mail
        s.sendmail(fromaddr, toaddr, text)
        # terminating the session
        
    s.quit()
  







  



